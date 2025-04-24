from ortools.sat.python import cp_model
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def load_data(file_path):
    xl = pd.ExcelFile(file_path)
    subjects_df = xl.parse('Mācību stundas', header=1)
    subjects_df = subjects_df[~subjects_df['Priekšmeti'].astype(str).str.strip().str.lower().eq('kopa')]
    subjects_df.dropna(subset=['Priekšmeti'], inplace=True)
    subjects_df.columns = subjects_df.columns.str.strip()

    teachers_df = xl.parse('Skolotāju grafiks', header=1)
    teachers_df = teachers_df[~teachers_df['Priekšmets'].astype(str).str.strip().str.lower().eq('kopa')]
    teachers_df.dropna(subset=['Priekšmets', 'Skolotāji'], inplace=True)
    teachers_df.columns = teachers_df.columns.str.strip()

    return subjects_df, teachers_df


def process_data(subjects_df, teachers_df):
    subject_hours = {}
    for _, row in subjects_df.iterrows():
        subject = row['Priekšmeti'].strip()
        if not subject or subject == 'Priekšmeti' or subject == 'Kopā':
            continue
        for class_id in subjects_df.columns[1:]:
            if pd.notna(row[class_id]):
                try:
                    hours = int(row[class_id])
                    if subject not in subject_hours:
                        subject_hours[subject] = {}
                    subject_hours[subject][class_id.strip()] = hours
                except ValueError:
                    print(f"Ошибка: некорректное значение часов для {subject} в классе {class_id}")

    teachers = {}
    teacher_subject_map = {}

    for _, row in teachers_df.iterrows():
        subject = row['Priekšmets'].strip()
        teacher = row['Skolotāji'].strip()

        classes = []
        if pd.notna(row['Klases']):
            classes = [x.strip() for x in str(row['Klases']).split('/') if x.strip()]

        max_hours = 40
        if pd.notna(row['Stundu skaits']):
            try:
                max_hours = int(row['Stundu skaits'])
            except ValueError:
                print(f"Ошибка: некорректная нагрузка для {teacher}")

        teachers[teacher] = {
            'subject': subject,
            'classes': classes,
            'max_hours_per_week': max_hours
        }

        if subject not in teacher_subject_map:
            teacher_subject_map[subject] = []
        teacher_subject_map[subject].append(teacher)

    return subject_hours, teachers, teacher_subject_map


def configure_model(subject_hours, teachers, teacher_subject_map):
    classes = list({class_id for subj in subject_hours.values() for class_id in subj})
    days = ['Pirmdiena', 'Otrdiena', 'Trešdiena', 'Ceturtdiena', 'Piektdiena']
    periods_per_day = 4
    weeks = 2

    for subject in subject_hours:
        if subject not in teacher_subject_map:
            print(f"Внимание: нет учителей для предмета {subject}!")

    def class_sort_key(class_id):
        # Извлекаем число и букву, например: '10A' → (10, 'A')
        num = int(''.join(filter(str.isdigit, class_id)))
        letter = ''.join(filter(str.isalpha, class_id))
        return (num, letter)

    classes.sort(key=class_sort_key)
    return classes, days, periods_per_day, weeks

def create_model(classes, days, periods_per_day, weeks, subject_hours, teachers, teacher_subject_map):
    model = cp_model.CpModel()
    subjects = list(subject_hours.keys())
    subject_index = {subj: idx for idx, subj in enumerate(subjects)}
    reverse_subject_index = {idx: subj for idx, subj in enumerate(subjects)}

    schedule = {}
    for week in range(weeks):
        for day in days:
            for period in range(periods_per_day):
                for class_id in classes:
                    schedule[(week, day, period, class_id)] = model.NewIntVar(
                        -1, len(subjects) - 1,
                        f'w{week}_d{day}_p{period}_c{class_id}')

    teacher_assign = {}
    for week in range(weeks):
        for day in days:
            for period in range(periods_per_day):
                for class_id in classes:
                    for subject in subjects:
                        if class_id in subject_hours.get(subject, {}):
                            for teacher in teacher_subject_map.get(subject, []):
                                if class_id in teachers[teacher]['classes']:
                                    key = (week, day, period, class_id, teacher)
                                    teacher_assign[key] = model.NewBoolVar(f't_{key}')

    for subject, class_data in subject_hours.items():
        subj_idx = subject_index[subject]
        for class_id, hours in class_data.items():
            slots = []
            for week in range(weeks):
                for day in days:
                    for period in range(periods_per_day):
                        slot = schedule[(week, day, period, class_id)]
                        is_this_subj = model.NewBoolVar(f'subj_{subject}_{week}_{day}_{period}_{class_id}')
                        model.Add(slot == subj_idx).OnlyEnforceIf(is_this_subj)
                        model.Add(slot != subj_idx).OnlyEnforceIf(is_this_subj.Not())
                        slots.append(is_this_subj)
            model.Add(sum(slots) == hours)

    for week in range(weeks):
        for day in days:
            for period in range(periods_per_day):
                for teacher in teachers:
                    teaching = []
                    for class_id in teachers[teacher]['classes']:
                        key = (week, day, period, class_id, teacher)
                        if key in teacher_assign:
                            teaching.append(teacher_assign[key])
                    if teaching:
                        model.Add(sum(teaching) <= 1)

    for teacher, data in teachers.items():
        total_lessons = []
        for week in range(weeks):
            for day in days:
                for period in range(periods_per_day):
                    for class_id in data['classes']:
                        key = (week, day, period, class_id, teacher)
                        if key in teacher_assign:
                            total_lessons.append(teacher_assign[key])
        if total_lessons:
            model.Add(sum(total_lessons) <= data['max_hours_per_week'] * weeks)

    for week in range(weeks):
        for day in days:
            for period in range(periods_per_day):
                for class_id in classes:
                    slot = schedule[(week, day, period, class_id)]
                    teachers_for_slot = []

                    for subject in subjects:
                        if class_id in subject_hours.get(subject, {}):
                            for teacher in teacher_subject_map.get(subject, []):
                                if class_id in teachers[teacher]['classes']:
                                    key = (week, day, period, class_id, teacher)
                                    if key in teacher_assign:
                                        model.Add(slot == subject_index[subject]).OnlyEnforceIf(teacher_assign[key])
                                        teachers_for_slot.append(teacher_assign[key])

                    if teachers_for_slot:
                        model.Add(sum(teachers_for_slot) == 1).OnlyEnforceIf(slot != -1)
                        model.Add(sum(teachers_for_slot) == 0).OnlyEnforceIf(slot == -1)

    for week in range(weeks):
        for day in days:
            for class_id in classes:
                for period in range(periods_per_day - 1):
                    current = schedule[(week, day, period, class_id)]
                    next_ = schedule[(week, day, period + 1, class_id)]

                    current_has_lesson = model.NewBoolVar(f'has_lesson_w{week}_d{day}_p{period}_c{class_id}')
                    next_empty = model.NewBoolVar(f'next_empty_w{week}_d{day}_p{period + 1}_c{class_id}')

                    model.Add(current != -1).OnlyEnforceIf(current_has_lesson)
                    model.Add(current == -1).OnlyEnforceIf(current_has_lesson.Not())

                    model.Add(next_ == -1).OnlyEnforceIf(next_empty)
                    model.Add(next_ != -1).OnlyEnforceIf(next_empty.Not())

                    model.AddBoolOr([current_has_lesson.Not(), next_empty.Not()])

    return model, schedule, teacher_assign, reverse_subject_index


def export_schedule_to_excel(schedule, teacher_assign, reverse_subject_index, classes, days, periods_per_day, weeks,
                             teachers, solver, file_path):
    wb = openpyxl.load_workbook(file_path)

    # Время уроков
    period_times = ["8:30-9:50", "10:00-11:20", "11:50-13:10", "13:15-14:35"]

    for week in range(weeks):
        sheet_title = f'Skolēnu stundu saraksts {week + 1}n.'
        if sheet_title in wb.sheetnames:
            ws = wb[sheet_title]
            wb.remove(ws)
        ws = wb.create_sheet(title=sheet_title)

        # Большой заголовок
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(classes))
        ws.cell(row=1, column=1).value = "Mācību priekšmetu stundu saraksts vispārējās vidējās izglītības programmai"
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Заголовки: Nr. | Stundas | классы...
        ws.cell(row=3, column=1).value = "Nr."
        ws.cell(row=3, column=2).value = "Stundas"
        for i, class_id in enumerate(classes):
            col = 3 + i
            ws.cell(row=3, column=col).value = class_id
            ws.cell(row=4, column=col).value = "Priekšmets"

        row_index = 5
        for day in days:
            ws.cell(row=row_index, column=1).value = day
            row_index += 1
            for period in range(periods_per_day):
                ws.cell(row=row_index, column=1).value = period + 1
                ws.cell(row=row_index, column=2).value = period_times[period]
                for i, class_id in enumerate(classes):
                    col = 3 + i
                    key = (week, day, period, class_id)
                    if key in schedule:
                        val = solver.Value(schedule[key])
                        if val != -1:
                            subject = reverse_subject_index[val]
                            ws.cell(row=row_index, column=col).value = subject
                row_index += 1

    wb.save(file_path)
    print(f'\n✅ Stundu saraksts saglabāts kā: {file_path}')

def _parse_time_slot(time_str):
    """Определяет номер периода по времени"""
    time_map = {
        "8:30-9:50": 0,
        "10:00-11:20": 1,
        "11:50-13:10": 2,
        "13:15-14:35": 3
    }
    for pattern, idx in time_map.items():
        if pattern in str(time_str):
            return idx
    return None

def solve_and_print(model, schedule, teacher_assign, reverse_subject_index, classes, days, periods_per_day, weeks, teachers):
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 600.0
    solver.parameters.num_search_workers = 8

    status = solver.Solve(model)

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        export_schedule_to_excel(
            schedule, teacher_assign, reverse_subject_index,
            classes, days, periods_per_day, weeks,
            teachers, solver, 'Schedule.xlsx'
        )
    else:
        print("❌ Решение не найдено.")


def main():
    subjects_df, teachers_df = load_data('Schedule.xlsx')
    subject_hours, teachers, teacher_subject_map = process_data(subjects_df, teachers_df)
    classes, days, periods_per_day, weeks = configure_model(subject_hours, teachers, teacher_subject_map)
    model, schedule, teacher_assign, reverse_subject_index = create_model(
        classes, days, periods_per_day, weeks,
        subject_hours, teachers, teacher_subject_map)
    solve_and_print(model, schedule, teacher_assign, reverse_subject_index,
                    classes, days, periods_per_day, weeks, teachers)


if __name__ == '__main__':
    main()

