import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model
import multiprocessing # only for multicore processing

# === PARAMETERS ===
file_path = "Schedule.xlsx"
days = ['Pirmdiena', 'Otrdiena', 'Trešdiena', 'Ceturtdiena', 'Piektdiena']
period_times = ["8:30-9:50", "10:00-11:20", "11:50-13:10", "13:15-14:35"]
periods_per_day = len(period_times)
weeks = 2
NUM_WORKERS = multiprocessing.cpu_count() # using all cors from CPU
time_limit = 1200

def load_data(file_path):
    xl = pd.ExcelFile(file_path)
    subjects_df = (
        xl.parse('Mācību stundas', header=1)
          .dropna(subset=['Priekšmeti'])
          .pipe(lambda df: df[df['Priekšmeti'].str.strip().str.lower() != 'kopā'])
    )
    subjects_df.columns = subjects_df.columns.str.strip()
    teachers_df = (
        xl.parse('Skolotāju grafiks', header=1)
          .dropna(subset=['Priekšmets','Skolotāji'])
    )
    teachers_df.columns = teachers_df.columns.str.strip()
    rooms_df = xl.parse('Kabinetu saraksts', header=1)
    rooms_df.columns = rooms_df.columns.str.strip()
    return subjects_df, teachers_df, rooms_df

def expand_class_range(entry, all_classes):
    for d in '/,;': entry = entry.replace(d, '/')
    out = []
    for tok in entry.split('/'):
        t = tok.strip()
        if not t: continue
        if t in all_classes:
            out.append(t)
        elif t.isdigit():
            out += [c for c in all_classes if c.startswith(t + '.')]
    return out

def prepare_data(subjects_df, teachers_df, rooms_df):
    class_list = subjects_df.columns[1:].tolist()
    # subject hours
    subject_hours = {}
    for _, row in subjects_df.iterrows():
        subj = row['Priekšmeti'].strip()
        for cls in class_list:
            if pd.notna(row[cls]):
                subject_hours.setdefault(subj, {})[cls] = int(row[cls])
    # teacher assignments + unavailability
    teacher_assignments = {}
    teacher_hours = {}
    teacher_unav = {}
    for _, row in teachers_df.iterrows():
        subj = row['Priekšmets'].strip()
        tchr = row['Skolotāji'].strip()
        classes = expand_class_range(str(row['Klases']), class_list)
        teacher_hours[tchr] = int(row.get('Stundu skaits', 40))
        slots = []
        for x in str(row.get('Nepieejamība','')).split(','):
            x = x.strip()
            if '-' in x:
                d,p,w = x.split('-')
                slots.append((d.strip(),int(p)-1,int(w)-1))
        teacher_unav[tchr] = slots
        for c in classes:
            teacher_assignments.setdefault((subj,c), []).append(tchr)
    # rooms
    room_map = {}
    all_rooms = []
    for _, row in rooms_df.iterrows():
        subj = row['Priekšmets'].strip()
        def split_clean(col):
            raw = row.get(col,'')
            if pd.isna(raw): return []
            return [s.strip() for s in str(raw).split(',') if s.strip()]
        gen = split_clean('Kabineti')
        lab = split_clean('Lab. Kabineti')
        spr = split_clean('Sporta zāle')
        all_rooms += gen+lab+spr
        if 'Sports' in subj or 'fizkultūra' in subj.lower():
            room_map[subj] = spr or gen
        else:
            room_map[subj] = gen+lab or spr
    for subj in subject_hours:
        if not room_map.get(subj):
            room_map[subj] = list(set(all_rooms))
    # valid subjects per class
    valid_subjects_by_class = {c:[] for c in class_list}
    for subj,cls_hours in subject_hours.items():
        for c in cls_hours:
            valid_subjects_by_class[c].append(subj)
    return subject_hours, teacher_assignments, room_map, teacher_hours, teacher_unav, class_list, valid_subjects_by_class

class SchoolScheduleModel:
    def __init__(self, classes, days, periods, weeks, subjects,
                 subject_hours, teachers, teacher_subject_map, rooms, unavailability, valid_subjects_by_class):
        m = cp_model.CpModel()
        self.model = m
        self.classes, self.days = classes, days
        self.periods, self.weeks = periods, weeks
        self.subjects = subjects
        self.subject_index = {s:i for i,s in enumerate(subjects)}
        self.subject_hours = subject_hours
        self.teachers = teachers
        self.teacher_subject_map = teacher_subject_map
        self.rooms = rooms
        # filter only valid unavailability
        self.unav = {
            t:[(d,p,w) for d,p,w in slots if d in days and 0<=p<periods and 0<=w<weeks]
            for t,slots in unavailability.items()
        }
        self.valid_subjects = valid_subjects_by_class

        self.all_rooms = sorted({r for lst in rooms.values() for r in lst})
        self.allowed_rooms_by_class = {
            c: sorted({self.all_rooms.index(r)
                       for subj in self.valid_subjects[c]
                       for r in self.rooms.get(subj, [])
                       if r in self.all_rooms})
            for c in self.classes
        }

        self.schedule = {}
        self.room_assign = {}
        self.teacher_assign = {}
        self.adj_vars = []
        self.balance_pen_vars = []
        self.teacher_gap_pen = []
        self.class_gap_pen = []
        self.room_change_pen = []
        self.penalty_consec = []

        self._create_variables()
        # symmetry break: first class first period nonempty
        m.Add(self.schedule[(0, days[0], 0, classes[0])] >= 0)
        self._add_constraints()
        self._add_balance()
        self._add_teacher_gaps()
        self._add_class_gaps()
        self._add_room_changes()
        self._add_unavailability()
        self._add_adjacency()
        m.Maximize(
            sum(self.adj_vars)
            - 2 * sum(self.balance_pen_vars)
            - 1 * sum(self.teacher_gap_pen)
            - 2 * sum(self.class_gap_pen)
            - 1 * sum(self.room_change_pen)
            - 2 * sum(self.penalty_consec)
        )

    def _create_variables(self):
        m = self.model
        for w in range(self.weeks):
            for d in self.days:
                for p in range(self.periods):
                    for c in self.classes:
                        vals = [-1]+[self.subject_index[s] for s in self.valid_subjects[c]]
                        self.schedule[(w,d,p,c)] = m.NewIntVarFromDomain(
                            cp_model.Domain.FromValues(vals),
                            f's_{w}_{d}_{p}_{c}'
                        )
                        dom = self.allowed_rooms_by_class[c]
                        if dom:
                            self.room_assign[(w,d,p,c)] = m.NewIntVarFromDomain(
                                cp_model.Domain.FromValues(dom),
                                f'r_{w}_{d}_{p}_{c}'
                            )
                        else:
                            self.room_assign[(w,d,p,c)] = m.NewIntVar(
                                0, len(self.all_rooms)-1, f'r_{w}_{d}_{p}_{c}'
                            )
                        for subj in self.valid_subjects[c]:
                            s_idx = self.subject_index[subj]
                            for t in self.teacher_subject_map.get(subj, []):
                                if (d,p,w) in self.unav.get(t, []):
                                    continue
                                key = (w,d,p,c,t)
                                tv = m.NewBoolVar(f't_{w}_{d}_{p}_{c}_{t}')
                                self.teacher_assign[key] = tv
                                sv = self.schedule[(w,d,p,c)]
                                m.Add(sv == s_idx).OnlyEnforceIf(tv)
                                m.Add(sv != s_idx).OnlyEnforceIf(tv.Not())

    def _add_constraints(self):
        m = self.model
        # exact hours per subject
        for s in self.subjects:
            idx = self.subject_index[s]
            for c,h in self.subject_hours.get(s,{}).items():
                flags=[]
                for w in range(self.weeks):
                    for d in self.days:
                        for p in range(self.periods):
                            b=m.NewBoolVar(f'flag_{s}_{c}_{w}_{d}_{p}')
                            sv=self.schedule[(w,d,p,c)]
                            m.Add(sv==idx).OnlyEnforceIf(b)
                            m.Add(sv!=idx).OnlyEnforceIf(b.Not())
                            flags.append(b)
                m.Add(sum(flags)==h)
        # teacher non overlap
        for w in range(self.weeks):
            for d in self.days:
                for p in range(self.periods):
                    for t in self.teachers:
                        occ=[self.teacher_assign[k]
                             for k in self.teacher_assign
                             if k[:3]==(w,d,p) and k[4]==t]
                        if occ:
                            m.Add(sum(occ)<=1)
        # teacher load
        for t,info in self.teachers.items():
            max_load=info['max_hours_per_week']*self.weeks
            all_vars=[v for k,v in self.teacher_assign.items() if k[4]==t]
            if all_vars:
                m.Add(sum(all_vars)<=max_load)
        # first period
        for w in range(self.weeks):
            for d in self.days:
                for c in self.classes:
                    m.Add(self.schedule[(w,d,0,c)]!=-1)
        # no empty spaces
        for w in range(self.weeks):
            for d in self.days:
                for c in self.classes:
                    for p in range(1,self.periods):
                        prev=self.schedule[(w,d,p-1,c)]
                        curr=self.schedule[(w,d,p,c)]
                        no_prev=m.NewBoolVar('')
                        has_curr=m.NewBoolVar('')
                        m.Add(prev==-1).OnlyEnforceIf(no_prev)
                        m.Add(prev!=-1).OnlyEnforceIf(no_prev.Not())
                        m.Add(curr!=-1).OnlyEnforceIf(has_curr)
                        m.Add(curr==-1).OnlyEnforceIf(has_curr.Not())
                        m.AddBoolOr([no_prev.Not(),has_curr.Not()])
        # room-subject compatibility
        for w in range(self.weeks):
            for d in self.days:
                for p in range(self.periods):
                    for c in self.classes:
                        sv=self.schedule[(w,d,p,c)]
                        rv=self.room_assign[(w,d,p,c)]
                        for subj,sidx in self.subject_index.items():
                            is_s=m.NewBoolVar('')
                            m.Add(sv==sidx).OnlyEnforceIf(is_s)
                            m.Add(sv!=sidx).OnlyEnforceIf(is_s.Not())
                            allowed=[i for i,r in enumerate(self.all_rooms) if r in self.rooms.get(subj,[])]
                            if allowed:
                                m.AddAllowedAssignments([rv],[[i] for i in allowed]).OnlyEnforceIf(is_s)
                            else:
                                m.AddAllowedAssignments([rv],[[i] for i in range(len(self.all_rooms))]).OnlyEnforceIf(is_s)
        # one class per room
        for w in range(self.weeks):
            for d in self.days:
                for p in range(self.periods):
                    for ridx in range(len(self.all_rooms)):
                        occ_bool=[]
                        for c in self.classes:
                            b=m.NewBoolVar(f'room_{ridx}_{w}_{d}_{p}_{c}')
                            m.Add(self.room_assign[(w,d,p,c)]==ridx).OnlyEnforceIf(b)
                            m.Add(self.room_assign[(w,d,p,c)]!=ridx).OnlyEnforceIf(b.Not())
                            occ_bool.append(b)
                        m.Add(sum(occ_bool)<=1)

        self.penalty_consec = []
        max_consecutive = 2
        for c in self.classes:
            for w in range(self.weeks):
                for d in self.days:
                    for start in range(self.periods - max_consecutive):
                        block = []
                        for i in range(max_consecutive + 1):
                            b = m.NewBoolVar(f'consec_{c}_{w}_{d}_{start + i}')
                            m.Add(self.schedule[(w, d, start + i, c)] != -1).OnlyEnforceIf(b)
                            m.Add(self.schedule[(w, d, start + i, c)] == -1).OnlyEnforceIf(b.Not())
                            block.append(b)
                        # penalty if maximum is exceeded
                        pen = m.NewBoolVar(f'pen_consec_{c}_{w}_{d}_{start}')
                        m.Add(sum(block) > max_consecutive).OnlyEnforceIf(pen)
                        m.Add(sum(block) <= max_consecutive).OnlyEnforceIf(pen.Not())
                        self.penalty_consec.append(pen)

    def _add_balance(self):
        m=self.model
        for subj in self.subjects:
            sidx=self.subject_index[subj]
            for c,h in self.subject_hours.get(subj,{}).items():
                base,extra=divmod(h,self.weeks)
                for w in range(self.weeks):
                    occ=[]
                    for d in self.days:
                        for p in range(self.periods):
                            b=m.NewBoolVar('')
                            m.Add(self.schedule[(w,d,p,c)]==sidx).OnlyEnforceIf(b)
                            m.Add(self.schedule[(w,d,p,c)]!=sidx).OnlyEnforceIf(b.Not())
                            occ.append(b)
                    low=m.NewBoolVar('')
                    m.Add(sum(occ)<base).OnlyEnforceIf(low)
                    m.Add(sum(occ)>=base).OnlyEnforceIf(low.Not())
                    hi=m.NewBoolVar('')
                    limit = base + (1 if w >= self.weeks - extra else 0)
                    m.Add(sum(occ)>limit).OnlyEnforceIf(hi)
                    m.Add(sum(occ)<=limit).OnlyEnforceIf(hi.Not())
                    self.balance_pen_vars += [low,hi]

    def _add_teacher_gaps(self):
        m=self.model
        for t,info in self.teachers.items():
            for w in range(self.weeks):
                for d in self.days:
                    teach=[]
                    for p in range(self.periods):
                        v=m.NewBoolVar('')
                        elems=[self.teacher_assign.get((w,d,p,c,t),m.NewConstant(0))
                               for c in info['classes']]
                        m.AddMaxEquality(v,elems)
                        teach.append(v)
                    for p in range(self.periods-2):
                        gap=m.NewBoolVar('')
                        m.AddBoolAnd([teach[p],teach[p+2],teach[p+1].Not()]).OnlyEnforceIf(gap)
                        m.AddBoolOr([teach[p].Not(),teach[p+2].Not(),teach[p+1]]).OnlyEnforceIf(gap.Not())
                        self.teacher_gap_pen.append(gap)

    def _add_class_gaps(self):
        m=self.model
        for c in self.classes:
            for w in range(self.weeks):
                for d in self.days:
                    cls=[]
                    for p in range(self.periods):
                        b=m.NewBoolVar('')
                        m.Add(self.schedule[(w,d,p,c)]!=-1).OnlyEnforceIf(b)
                        m.Add(self.schedule[(w,d,p,c)]== -1).OnlyEnforceIf(b.Not())
                        cls.append(b)
                    for p in range(self.periods-2):
                        gap=m.NewBoolVar('')
                        m.AddBoolAnd([cls[p],cls[p+2],cls[p+1].Not()]).OnlyEnforceIf(gap)
                        m.AddBoolOr([cls[p].Not(),cls[p+2].Not(),cls[p+1]]).OnlyEnforceIf(gap.Not())
                        self.class_gap_pen.append(gap)

    def _add_room_changes(self):
        m=self.model
        for w in range(self.weeks):
            for d in self.days:
                for c in self.classes:
                    for p in range(self.periods-1):
                        s1=self.schedule[(w,d,p,c)]
                        s2=self.schedule[(w,d,p+1,c)]
                        same_s=m.NewBoolVar('')
                        m.Add(s1==s2).OnlyEnforceIf(same_s)
                        m.Add(s1!=s2).OnlyEnforceIf(same_s.Not())
                        r1=self.room_assign[(w,d,p,c)]
                        r2=self.room_assign[(w,d,p+1,c)]
                        same_r=m.NewBoolVar('')
                        m.Add(r1==r2).OnlyEnforceIf(same_r)
                        m.Add(r1!=r2).OnlyEnforceIf(same_r.Not())
                        change=m.NewBoolVar('')
                        m.AddBoolAnd([same_s,same_r.Not()]).OnlyEnforceIf(change)
                        m.AddBoolOr([same_s.Not(),same_r]).OnlyEnforceIf(change.Not())
                        self.room_change_pen.append(change)

    def _add_unavailability(self):
        m=self.model
        for (w,d,p,c,t),var in self.teacher_assign.items():
            if (d,p,w) in self.unav.get(t,[]):
                m.Add(var==0)

    def _add_adjacency(self):
        m = self.model
        for w in range(self.weeks):
            for d in self.days:
                for c in self.classes:
                    for p in range(self.periods - 1):
                        v1 = self.schedule[(w, d, p, c)]
                        v2 = self.schedule[(w, d, p + 1, c)]
                        b = m.NewBoolVar(f"adj_{w}_{d}_{p}_{c}")
                        # if b then v1==v2 and both non-empty
                        m.Add(v1 == v2).OnlyEnforceIf(b)
                        m.Add(v1 != v2).OnlyEnforceIf(b.Not())
                        m.Add(v1 >= 0).OnlyEnforceIf(b)
                        m.Add(v2 >= 0).OnlyEnforceIf(b)
                        self.adj_vars.append(b)

    def solve(self, time_limit):
        solver=cp_model.CpSolver()
        solver.parameters.num_search_workers=NUM_WORKERS
        solver.parameters.cp_model_presolve=True
        solver.parameters.search_branching=cp_model.FIXED_SEARCH
        solver.parameters.max_time_in_seconds=time_limit*0.2
        status=solver.Solve(self.model)
        if status==cp_model.OPTIMAL:
            return status,solver
        solver=cp_model.CpSolver()
        solver.parameters.num_search_workers=NUM_WORKERS
        solver.parameters.cp_model_presolve=True
        solver.parameters.search_branching=cp_model.PORTFOLIO_SEARCH
        solver.parameters.max_time_in_seconds=time_limit*0.8
        solver.parameters.log_search_progress=True
        status=solver.Solve(self.model)
        return status,solver


def style_worksheet(ws):
    thin=Border(left=Side('thin'),right=Side('thin'),
                top=Side('thin'),bottom=Side('thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border=thin
    for i in range(1,ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].width=18

def export_schedule(df,wb):
    for w in [1,2]:
        sheet=f"Skolēnu stundu saraksts {w}n."
        if sheet in wb.sheetnames: wb.remove(wb[sheet])
        ws=wb.create_sheet(sheet)
        classes=sorted(df['Class'].unique())
        ws.cell(2,1).value="Nr."; ws.cell(2,2).value="Stundas"
        for i,cls in enumerate(classes): ws.cell(2,3+i).value=cls
        row=4; week_label=f"{w}. nedēļa"
        for day in days:
            ws.cell(row,1).value=day; row+=1
            for p in range(periods_per_day):
                ws.cell(row,1).value=p+1
                ws.cell(row,2).value=period_times[p]
                mask=(df['Week']==week_label)&(df['Day']==day)&(df['Period']==p+1)
                slice_=df[mask]
                for i,cls in enumerate(classes):
                    sub=slice_[slice_['Class']==cls]
                    if not sub.empty:
                        subj=sub.iloc[0]['Subject']
                        room=sub.iloc[0]['Room']
                        ws.cell(row,3+i).value=f"{subj}\n({room})"
                row+=1
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=2+len(classes))
        title=ws.cell(1,1)
        title.value=f"Skolēnu stundu saraksts ({w}. nedēļa)"
        title.font=Font(size=14,bold=True)
        title.alignment=Alignment(horizontal="center",vertical="center")
        style_worksheet(ws)

def export_teacher_timetables(df,wb):
    for w in [1,2]:
        sheet=f"Skolotāju stundu saraksts {w}n."
        if sheet in wb.sheetnames: wb.remove(wb[sheet])
        ws=wb.create_sheet(sheet)
        classes=sorted(df['Class'].unique())
        ws.cell(2,1).value="Nr."; ws.cell(2,2).value="Stundas"
        for i,cls in enumerate(classes): ws.cell(2,3+i).value=cls
        row=4; week_label=f"{w}. nedēļa"
        for day in days:
            ws.cell(row,1).value=day; row+=1
            for p in range(periods_per_day):
                ws.cell(row,1).value=p+1
                ws.cell(row,2).value=period_times[p]
                mask=(df['Week']==week_label)&(df['Day']==day)&(df['Period']==p+1)
                slice_=df[mask]
                for i,cls in enumerate(classes):
                    sub=slice_[slice_['Class']==cls]
                    if not sub.empty:
                        teacher=sub.iloc[0]['Teacher']
                        room=sub.iloc[0]['Room']
                        ws.cell(row,3+i).value=f"{teacher}\n({room})"
                row+=1
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=2+len(classes))
        title=ws.cell(1,1)
        title.value=f"Skolotāju stundu saraksts ({w}. nedēļa)"
        title.font=Font(size=14,bold=True)
        title.alignment=Alignment(horizontal="center",vertical="center")
        style_worksheet(ws)

def export_schedule_from_model(model,solver,file_path):
    data=[]
    for (w,d,p,c),var in model.schedule.items():
        idx=solver.Value(var)
        if idx<0: continue
        subj=model.subjects[idx]
        room=model.all_rooms[solver.Value(model.room_assign[(w,d,p,c)])]
        teacher=next((t for (ww,dd,pp,cc,t),v in model.teacher_assign.items()
                      if (ww,dd,pp,cc)==(w,d,p,c) and solver.Value(v)==1),None)
        if teacher is None:
            poss=model.teacher_subject_map.get(subj,[])
            teacher=poss[0] if poss else ""
        data.append({
            'Week':f"{w+1}. nedēļa",
            'Day':d,
            'Period':p+1,
            'Time':period_times[p],
            'Class':c,
            'Subject':subj,
            'Teacher':teacher,
            'Room':room
        })
    df=pd.DataFrame(data)
    wb=openpyxl.load_workbook(file_path)
    export_schedule(df,wb)
    export_teacher_timetables(df,wb)
    wb.save(file_path)

if __name__=="__main__":
    subj_df, teach_df, rooms_df = load_data(file_path)
    subj_hours, teach_assigns, room_map, teach_hours, teach_unav, classes, valid_subjects = prepare_data(
        subj_df, teach_df, rooms_df)
    teachers_struct={}
    for (s,c),lst in teach_assigns.items():
        for t in lst:
            teachers_struct.setdefault(t,{'classes':set(),'max_hours_per_week':teach_hours[t]})
            teachers_struct[t]['classes'].add(c)
    for t in teachers_struct:
        teachers_struct[t]['classes']=list(teachers_struct[t]['classes'])
    teacher_subject_map={}
    for (s,c),lst in teach_assigns.items():
        teacher_subject_map.setdefault(s,set()).update(lst)
    teacher_subject_map={k:list(v) for k,v in teacher_subject_map.items()}
    model=SchoolScheduleModel(
        classes=classes,
        days=days,
        periods=periods_per_day,
        weeks=weeks,
        subjects=list(subj_hours.keys()),
        subject_hours=subj_hours,
        teachers=teachers_struct,
        teacher_subject_map=teacher_subject_map,
        rooms=room_map,
        unavailability=teach_unav,
        valid_subjects_by_class=valid_subjects
    )

    status, solver = model.solve(time_limit)
    print(f"Objective = {solver.ObjectiveValue():.1f}, best bound = {solver.BestObjectiveBound():.1f}")
    if status == cp_model.OPTIMAL:
        print("Optimal solution!")
    elif status == cp_model.FEASIBLE:
        print("Feasible solution:")
    else:
        print("No solution available.")
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        export_schedule_from_model(model,solver,file_path)
        print("Schedule saved.")